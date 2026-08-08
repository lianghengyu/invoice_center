import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import SAVED_FOLDER

HEADERS = [
    ('文件名', 20),
    ('发票类型', 18),
    ('发票代码', 16),
    ('发票号码', 16),
    ('开票日期', 16),
    ('购买方名称', 30),
    ('购买方税号', 24),
    ('销售方名称', 30),
    ('销售方税号', 24),
    ('金额', 14),
    ('税额', 14),
    ('价税合计', 14),
    ('校验码', 26),
]

FIELD_KEYS = [
    'filename', 'invoice_type', 'invoice_code', 'invoice_number',
    'invoice_date', 'buyer_name', 'buyer_tax_id', 'seller_name',
    'seller_tax_id', 'amount', 'tax_amount', 'total_amount', 'check_code',
]

DEFAULT_FILENAME = '发票识别结果.xlsx'

HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CELL_FONT = Font(name='微软雅黑', size=10)
CELL_ALIGN = Alignment(vertical='center', wrap_text=True)


def _apply_header_style(ws):
    for col_idx, (header, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = 'A2'


def _apply_cell_style(cell):
    cell.font = CELL_FONT
    cell.alignment = CELL_ALIGN
    cell.border = THIN_BORDER


def _build_existing_map(ws):
    filename_to_row = {}
    for row in range(2, ws.max_row + 1):
        existing_name = ws.cell(row=row, column=1).value
        if existing_name:
            filename_to_row[existing_name] = row
    return filename_to_row


def _write_row(ws, row_idx, item):
    for col_idx, key in enumerate(FIELD_KEYS, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, ''))
        _apply_cell_style(cell)


def export_to_excel(results, filename=DEFAULT_FILENAME):
    filepath = os.path.join(SAVED_FOLDER, filename)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        filename_to_row = _build_existing_map(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '识别结果'
        _apply_header_style(ws)
        filename_to_row = {}

    next_row = ws.max_row + 1 if ws.max_row >= 1 else 2

    for item in results:
        name = item.get('filename', '')
        if name and name in filename_to_row:
            _write_row(ws, filename_to_row[name], item)
        else:
            _write_row(ws, next_row, item)
            if name:
                filename_to_row[name] = next_row
            next_row += 1

    wb.save(filepath)
    return filepath
