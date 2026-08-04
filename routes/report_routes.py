import os
import uuid
import time
from flask import Blueprint, request, jsonify, send_file, after_this_request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import UPLOAD_FOLDER, EXPORT_FOLDER
from services.report_service import (
    get_overview_stats, get_field_recognition_rates, get_invoice_type_distribution,
    get_daily_trend, get_engine_comparison, get_annotation_comparison,
    get_batch_list, save_annotations, export_report_data, TRACKED_FIELDS,
)

report_bp = Blueprint('report', __name__, url_prefix='/api/report')

FIELD_LABELS = {
    'invoice_type': '发票类型',
    'invoice_code': '发票代码',
    'invoice_number': '发票号码',
    'invoice_date': '开票日期',
    'buyer_name': '购买方名称',
    'buyer_tax_id': '购买方税号',
    'seller_name': '销售方名称',
    'seller_tax_id': '销售方税号',
    'amount': '金额',
    'tax_amount': '税额',
    'total_amount': '价税合计',
    'check_code': '校验码',
}


@report_bp.route('/stats', methods=['GET'])
def get_stats():
    overview = get_overview_stats()
    field_rates = get_field_recognition_rates()
    type_dist = get_invoice_type_distribution()
    trend_days = request.args.get('days', 30, type=int)
    trend = get_daily_trend(trend_days)
    engine_cmp = get_engine_comparison()

    return jsonify({
        'success': True,
        'overview': overview,
        'field_rates': field_rates,
        'field_labels': FIELD_LABELS,
        'type_distribution': type_dist,
        'daily_trend': trend,
        'engine_comparison': engine_cmp,
    })


@report_bp.route('/batches', methods=['GET'])
def list_batches():
    limit = request.args.get('limit', 50, type=int)
    batches = get_batch_list(limit)
    return jsonify({'success': True, 'batches': batches})


@report_bp.route('/annotation/upload', methods=['POST'])
def upload_annotation():
    f = request.files.get('file')
    batch_id = request.form.get('batch_id', '')

    if not f or not f.filename:
        return jsonify({'success': False, 'message': '未上传文件'}), 400

    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': '仅支持 .xlsx 格式'}), 400

    if not batch_id:
        return jsonify({'success': False, 'message': '请选择要对比的批次'}), 400

    saved_path = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4().hex}.xlsx")
    f.save(saved_path)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(saved_path, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {}
        known_headers = {
            '文件名': 'filename', '发票类型': 'invoice_type', '发票代码': 'invoice_code',
            '发票号码': 'invoice_number', '开票日期': 'invoice_date',
            '购买方名称': 'buyer_name', '购买方税号': 'buyer_tax_id',
            '销售方名称': 'seller_name', '销售方税号': 'seller_tax_id',
            '金额': 'amount', '税额': 'tax_amount', '价税合计': 'total_amount',
            '校验码': 'check_code',
        }
        for idx, h in enumerate(headers):
            if h and h.strip() in known_headers:
                header_map[idx] = known_headers[h.strip()]

        if 'filename' not in header_map.values():
            return jsonify({'success': False, 'message': 'Excel 中需包含"文件名"列'}), 400

        annotations = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for idx, field in header_map.items():
                val = row[idx] if idx < len(row) else ''
                record[field] = str(val).strip() if val else ''
            if record.get('filename'):
                annotations.append(record)

        wb.close()

        if not annotations:
            return jsonify({'success': False, 'message': 'Excel 中无有效标注数据'}), 400

        save_annotations(batch_id, annotations)
        comparison = get_annotation_comparison(batch_id)

        return jsonify({
            'success': True,
            'message': f'已导入 {len(annotations)} 条标注数据',
            'comparison': comparison,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'解析失败: {str(e)}'}), 400
    finally:
        if os.path.exists(saved_path):
            os.remove(saved_path)


@report_bp.route('/annotation/compare', methods=['GET'])
def annotation_compare():
    batch_id = request.args.get('batch_id', '')
    comparison = get_annotation_comparison(batch_id or None)
    return jsonify({'success': True, 'comparison': comparison})


@report_bp.route('/export', methods=['GET'])
def export_report():
    data = export_report_data()

    wb = Workbook()

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws, headers):
        for col, (name, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width
        ws.freeze_panes = 'A2'

    # Sheet 1: 总览
    ws1 = wb.active
    ws1.title = '总览'
    overview = data['overview']
    stats_items = [
        ('总识别图片数', overview['total_images']),
        ('识别成功数', overview['success_count']),
        ('识别成功率', f"{overview['success_rate']}%"),
        ('平均耗时(秒)', overview['avg_duration']),
        ('批次总数', overview['batch_count']),
    ]
    style_header(ws1, [('指标', 20), ('数值', 15)])
    for row, (label, val) in enumerate(stats_items, 2):
        ws1.cell(row=row, column=1, value=label).border = thin_border
        ws1.cell(row=row, column=2, value=val).border = thin_border

    # Sheet 2: 字段识别率
    ws2 = wb.create_sheet('字段识别率')
    style_header(ws2, [('字段', 20), ('识别率(%)', 15)])
    for row, field in enumerate(TRACKED_FIELDS, 2):
        ws2.cell(row=row, column=1, value=FIELD_LABELS.get(field, field)).border = thin_border
        ws2.cell(row=row, column=2, value=data['field_rates'].get(field, 0)).border = thin_border

    # Sheet 3: 发票类型分布
    ws3 = wb.create_sheet('发票类型分布')
    style_header(ws3, [('发票类型', 30), ('数量', 12)])
    for row, item in enumerate(data['type_distribution'], 2):
        ws3.cell(row=row, column=1, value=item['name']).border = thin_border
        ws3.cell(row=row, column=2, value=item['value']).border = thin_border

    # Sheet 4: 每日趋势
    ws4 = wb.create_sheet('每日趋势')
    style_header(ws4, [('日期', 14), ('总数', 10), ('成功数', 10), ('成功率(%)', 12)])
    for row, item in enumerate(data['daily_trend'], 2):
        ws4.cell(row=row, column=1, value=item['date']).border = thin_border
        ws4.cell(row=row, column=2, value=item['total']).border = thin_border
        ws4.cell(row=row, column=3, value=item['success']).border = thin_border
        rate = round(item['success'] / item['total'] * 100, 1) if item['total'] > 0 else 0
        ws4.cell(row=row, column=4, value=rate).border = thin_border

    # Sheet 5: 识别明细
    ws5 = wb.create_sheet('识别明细')
    detail_headers = [
        ('文件名', 20), ('是否成功', 10), ('耗时(s)', 10), ('OCR引擎', 12), ('检测模型', 12),
        ('发票类型', 18), ('发票代码', 16), ('发票号码', 16), ('开票日期', 14),
        ('购买方名称', 25), ('购买方税号', 22), ('销售方名称', 25), ('销售方税号', 22),
        ('金额', 12), ('税额', 12), ('价税合计', 12), ('校验码', 24), ('识别时间', 18),
    ]
    style_header(ws5, detail_headers)
    for row, rec in enumerate(data['records'], 2):
        vals = [
            rec.get('filename', ''),
            '成功' if rec.get('is_success') else '失败',
            rec.get('duration', ''),
            rec.get('ocr_engine', ''),
            rec.get('detector_model', ''),
            rec.get('invoice_type', ''),
            rec.get('invoice_code', ''),
            rec.get('invoice_number', ''),
            rec.get('invoice_date', ''),
            rec.get('buyer_name', ''),
            rec.get('buyer_tax_id', ''),
            rec.get('seller_name', ''),
            rec.get('seller_tax_id', ''),
            rec.get('amount', ''),
            rec.get('tax_amount', ''),
            rec.get('total_amount', ''),
            rec.get('check_code', ''),
            rec.get('created_at', ''),
        ]
        for col, val in enumerate(vals, 1):
            ws5.cell(row=row, column=col, value=val).border = thin_border

    export_name = f"识别报表_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_FOLDER, export_name)
    wb.save(filepath)

    @after_this_request
    def cleanup(response):
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except OSError:
            pass
        return response

    return send_file(filepath, as_attachment=True, download_name=export_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
